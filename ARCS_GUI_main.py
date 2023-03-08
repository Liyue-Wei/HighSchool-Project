import ttkbootstrap as ttk
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import time
import os
import pyautogui
import threading
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl

os.popen('start chrome --remote-debugging-port=9527 --user-data-dir="C:\chromedriver_win32"')
opt = Options()
opt.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
driver = webdriver.Chrome(options=opt)
driver.minimize_window()

(width, height) = pyautogui.size()
if (width>1920 and height>1080):
    screensize = ("1280x650")

else:
    screensize = ("1000x520")

def set_interval(func, sec):
    def func_wrapper():
        set_interval(func, sec)
        func()
    
    timer_set = threading.Timer(sec, func_wrapper)
    timer_set.start()
    return timer_set

def time_set():
    time_set = time.localtime()
    time_hour = time_set.tm_hour
    time_min = time_set.tm_min
    time_sec = time_set.tm_sec
    time_wday = (time_set.tm_wday+1)
    time_output = (str(time_hour)+str(time_min))
  
    schedule = [str(810), str(910), str(1020), str(1120), str(1320), str(1420), str(1520)]
    schedule_2 = [str(90), str(100), str(1110), str(1210), str(1410), str(1510), str(1610)]
    
    for section in range(0, len(schedule)):
        if (int(schedule[section]) == int(time_output) and time_sec==0):
            # print(section, "\n")
            # print(time_output, "\n")
            attend_class(section+1)

        elif int(schedule_2[section]) == int(time_output):
            messagebox.showinfo("下課時間")
            driver.get("https://google.com/")
            driver.minimize_window
    
    time_now.set(time.strftime("%Y/%m/%d %H:%M:%S", time.localtime())) 

def time_set_2():  
    time_set = time.localtime()
    time_hour = time_set.tm_hour
    time_min = time_set.tm_min
    time_output = (str(time_hour)+str(time_min))

    schedule = [str(810), str(910), str(1020), str(1120), str(1320), str(1420), str(1520)]
    schedule_2 = [str(900), str(1000), str(1110), str(1210), str(1410), str(1510), str(1610)] 

    for section in range(0, len(schedule)): 
        print(time_output)
        print(int(schedule[section]))
        print(int(schedule_2[section]))

        if (int(schedule[section]) < int(time_output) and int(time_output) < int(schedule_2[section])):
            attend_class(section)
            break
        
        else:
            print("False"+'\n')

win = ttk.Window(themename="cerculean")
win.geometry(screensize)
win.resizable(0, 0)
win.title("ARCS By K3S12")

def path():
    file_path = os.path.abspath(os.path.dirname(__file__))
    length = len(file_path)
    file_path_2 = []
    for i in range(0, length):       
        if file_path[i] == "\\":
            file_path_2.append("/")

        else :
            file_path_2.append(file_path[i])

    return (file_path_2)

def path_function(direct_path):
    file_path = path()
    file_path.append(direct_path)
    file_path = "".join(file_path)
    return(file_path)

def attend_class(section):
    auto_login(section)

def login():
    url = "https://accounts.google.com/v3/signin/identifier?dsh=S-738053007%3A1669617430712505&continue=https%3A%2F%2Fwww.google.com.tw%2Fwebhp%3Ftab%3Drw&ec=GAZAmgQ&hl=zh-TW&passive=true&flowName=GlifWebSignIn&flowEntry=ServiceLogin&ifkv=ARgdvAsS_WSw2KMw9woOHy9hfncKG2YsKx5sKtMqGDbjT1b3cfMqDdWGdIRfwi3fjpKnpmJYALlmUA"
    driver.get(url)
    driver.maximize_window()

def menu():   
    direct_path = ("/User_DATA/ARCS_menu.pyw")
    file_path = path_function(direct_path)
    os.system(file_path)

def guide():
    direct_path = ("/User_DATA/ARCS_User's_Guide.docx")
    file_path = path_function(direct_path)
    os.system(file_path)

def Eng_Ver():
    direct_path = ("/ARCS_GUI_Eng.pyw")
    file_path = path_function(direct_path)
    os.system(file_path)

def win_close():
    driver.minimize_window()
    driver.maximize_window()
    pyautogui.hotkey("ctrl", "shift", "w")
    os._exit(False)

def schedule():
    direct_path = ("/User_DATA/schedule.xlsx")
    file_path = path_function(direct_path)
    os.system(file_path)

def links():
    direct_path = ("/User_DATA/links.xlsx")
    file_path = path_function(direct_path)
    os.system(file_path)

def auto_login(section):
    print(section)
    direct_path = "/User_DATA/schedule.xlsx"
    workbook = openpyxl.load_workbook(path_function(direct_path))
    sheet = workbook.worksheets[0]
    class_list = []
    time_wday = (time.localtime().tm_wday)+1
    print(time_wday)

    for i in range(0, sheet.max_row):
        class_list.append(sheet.cell(row = i+1, column = time_wday).value)

    # print(class_list)
    # print(section-1)
    class_now = str(class_list[section-1])
    print(class_now)
    messagebox.showinfo("現在的課程", "現在是"+class_now+"課")

    direct_path_2 = "/User_DATA/links.xlsx"
    workbook_2 = openpyxl.load_workbook(path_function(direct_path_2))
    sheet_2 = workbook_2.worksheets[0]
    class_list_2 = []
    time_wday = (time.localtime().tm_wday)+1

    for i in range(0, sheet_2.max_row):
        class_list_2.append(sheet_2.cell(row = i+1, column = 1).value)

    for i in range(0, len(class_list_2)):
        if str(class_list_2[i]) == class_now:
            class_link = str(sheet_2.cell(row = i+1, column = 2).value)
    
    url = "https://meet.google.com/"
    print(url+class_link)
    driver.get(url+"/"+class_link)
    driver.maximize_window()

class GUI_interface:
    global time_now
    time_now = tk.StringVar()
    pic_path = path()
    pic_path.append("/User_DATA/K3S12.png")
    pic_path = "".join(pic_path)
    pic_logo = Image.open(pic_path)
    pic = ImageTk.PhotoImage(pic_logo)
    set_interval(time_set, 1)
    
    tk.Label(win, text="線上課程自動點名系統", font=("微軟正黑體", 24)).pack(side=tk.TOP, anchor=tk.NW, padx=10, pady=5)
    tk.Label(win, text="Version 1.0", font=("微軟正黑體", 12)).pack(anchor=tk.NE, padx=10, pady=0)
    tk.Label(win, image=pic).pack(side=tk.BOTTOM, anchor=tk.SE, padx=15, pady=15)
    tk.Label(win, textvariable=time_now, font=("微軟正黑體", 16)).pack(anchor=tk.SE, padx=10, pady=0)

    tk.Button(win, text="Attend Class 上課", font=("微軟正黑體", 16), command=time_set_2).pack(anchor=tk.W, fill=tk.BOTH, padx=10, pady=10, ipady=10)
    tk.Button(win, text="Quick Links to Current Courses 課程快速連結 (Google Classroom)", font=("微軟正黑體", 16)).pack(anchor=tk.W, fill=tk.BOTH, padx=10, pady=10, ipady=10)
    
    tk.Label(win, text=" ", font=("微軟正黑體", 16)).pack(anchor=tk.N, padx=10, pady=5, fill=tk.X)  
    tk.Label(win, text="編輯與選項", font=("微軟正黑體", 14)).pack(anchor=tk.W, padx=10, pady=0, ipady=10)
    tk.Button(win, text="編輯連結", font=("微軟正黑體", 14), command=links).pack(side=tk.LEFT, anchor=tk.N, padx=10, pady=10, ipady=5)
    tk.Button(win, text="編輯課表", font=("微軟正黑體", 14), command=schedule).pack(side=tk.LEFT, anchor=tk.N, padx=10, pady=10, ipady=5)
    tk.Button(win, text="首次登入", font=("微軟正黑體", 14), command=login).pack(side=tk.LEFT, anchor=tk.N, padx=10, pady=10, ipady=5)
    tk.Button(win, text="使用指南", font=("微軟正黑體", 14), command=guide).pack(side=tk.LEFT, anchor=tk.N, padx=10, pady=10, ipady=5)
    # tk.Button(win, text="English Version", font=("微軟正黑體", 14), command=Eng_Ver).pack(side=tk.LEFT, anchor=tk.N, padx=10, pady=10, ipady=5)
    tk.Button(win, text="(新功能)", font=("微軟正黑體", 14)).pack(side=tk.LEFT, anchor=tk.N, padx=10, pady=10, ipady=5)
    # tk.Button(win, text="功能列表", font=("微軟正黑體", 14), command=menu).pack(side=tk.LEFT, anchor=tk.N, padx=10, pady=10, ipady=5)
    tk.Button(win, text="(新功能)", font=("微軟正黑體", 14)).pack(side=tk.LEFT, anchor=tk.N, padx=10, pady=10, ipady=5)
    tk.Button(win, text="關閉", font=("微軟正黑體", 14), command=win_close).pack(anchor=tk.N, padx=10, pady=10, fill=tk.X, ipady=5)
   
win.mainloop()
